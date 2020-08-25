"""
analyze top log and output csv file (row: time, column: process Id and command name)
    option:
        --withExcel : Output Excel File(include line graph) and csv file
"""

import glob
import sys
import csv
import openpyxl
from typing import List, Dict
from openpyxl.chart import LineChart, Reference, Series
import matplotlib.pyplot as plt

# Fixed value
PID_INDEX = 0
CPU_INDEX = 8
MEM_INDEX = 9
COMMAND_INDEX = 11
PID_VALUE_COLUMN_COUNT = 12
EXCEL_OPTION = '--withExcel'
VIEW_GRAPH_OPTION = '--viewGraph'

# Changeable values
FILTER_VALUE = 1.0  # Output only values more than this value
RANK_TOP_LIMIT = 5  # Maximum number of processes to output
OUTPUT_TOP_MEM_NAME = 'top_memory_result'   # Output file name and graph title
OUTPUT_TOP_CPU_NAME = 'top_cpu_result'      # Output file name and graph title


def view_line_graph(id: int, name: str, header: List[str], big_order_indexes: List[int], array2d: List[List[str]]):
    times = [array2d[i][0] for i in range(len(array2d)) if i < len(array2d) - 1]
    x_alias = [i for i in times[::int(len(times) / 10)]]
    fig, axes = plt.subplots()
    for col in range(len(big_order_indexes)):
        if col == 0:
            # skip because column 0 is time
            continue
        if col > RANK_TOP_LIMIT or col > len(big_order_indexes):
            break
        y_values = []
        for row in range(len(array2d)):
            if not row < len(array2d) - 1:
                break
            y_values.append(
                float(array2d[row][big_order_indexes[col]]) if array2d[row][big_order_indexes[col]] != '' else None)
        axes.plot(times, y_values, label=header[col])
    axes.set_title(name)
    axes.set_xlabel('Time')
    axes.set_xticks(x_alias)
    axes.set_ylabel('Use Rate [%]')
    axes.set_ylim(0, 100)
    axes.legend()
    axes.grid()
    plt.xticks(rotation=40)


def create_excel_line_graph(sheet, rows_num: int):
    """
    Discription:
        undone function (Because can't check the operation in the developer environment).
    :param sheet:
    :param rows_num:
    :return: void
    """
    chart = openpyxl.chart.LineChart()
    chart.title = 'プロセス毎の使用率'
    chart.x_axis.title = "時刻"
    chart.y_axis.title = '使用量 [%]'
    chart.height = 16
    chart.width = 24
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 100
    time_refs = Reference(sheet, min_col=1, max_col=1, min_row=2, max_row=rows_num - 1)
    pid_refs = Reference(sheet, min_col=2, max_col=RANK_TOP_LIMIT, min_row=1, max_row=rows_num - 1)
    chart.add_data(pid_refs, titles_from_data=True)
    chart.set_categories(time_refs)
    sheet.add_chart(chart, "A" + str(rows_num + 3))


def write_excel_file(filename, header, big_order_indexes, array2d):
    """

    :param filename:
    :param header:
    :param big_order_indexes:
    :param array2d:
    :return: void
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'result'
    sheet = wb['result']
    for i in range(len(header)):
        sheet.cell(row=1, column=i + 1).value = header[i]
    for row in range(len(array2d)):
        for col in range(len(big_order_indexes)):
            sheet.cell(row=row + 2, column=col + 1).value = array2d[row][big_order_indexes[col]]
    create_excel_line_graph(sheet, len(array2d))
    wb.save('../output/' + filename + '.xlsx')


def write_csv_file(filename: str, header: List[str], max_order_indexes: List[int], array2d: List[List[str]]):
    """

    :param filename:
    :param header:
    :param max_order_indexes:
    :param array2d:
    :return: void
    """
    with open('../output/' + filename + '.csv', 'w', newline='') as csvfile:
        writer = csv.writer(csvfile, lineterminator='\n')
        writer.writerow(header)
        # writer.writerows(array2d)
        for row in range(len(array2d)):
            writer.writerow([array2d[row][index] for index in max_order_indexes])


def create_max_value_order(str_max_values: List[str]):
    """

    :param str_max_values:
    :return: Index list when sorted by maximum values per process id
    """
    max_values: List[float] = list(map(lambda s: float(s), str_max_values))
    sorted_max_values = sorted(max_values, reverse=True)
    big_order_indexes: List[int] = [max_values.index(v) + 1 for v in sorted_max_values]
    return [0] + big_order_indexes


def create_max_value_per_pid(array2d: List[List[str]]):
    """

    :param array2d:
    :return: void
    """
    rows_num: int = len(array2d)
    columns_num: int = len(array2d[0])
    max_values: List[str] = []
    for j in range(columns_num):
        if j == 0:
            continue
        max = '0.0'
        for i in range(rows_num):
            if array2d[i][j] != '' and float(max) < float(array2d[i][j]):
                max = array2d[i][j]
        max_values.append(max)
    return max_values


def pack_empty_str(pid_count: int, array2d: List[List[str]]):
    """

    :param pid_count:
    :param array2d:
    :return: void
    """
    for record in array2d:
        diff = pid_count - (len(record) - 1)
        record.extend([''] * diff)


def write_file_and_view_graph(id: int, name: str, pids: List[str], array2d: List[List[str]], is_output_excel: bool,
                              is_view_graph: bool):
    """

    :param name:
    :param pids:
    :param array2d:
    :param is_output_excel:
    :param is_view_graph:
    :return: void
    """
    pack_empty_str(len(pids), array2d)
    max_values_per_pid: List[str] = create_max_value_per_pid(array2d)
    big_order_indexes: List[int] = create_max_value_order(max_values_per_pid)
    header = [''] + pids
    array2d.append(['MAX:'] + max_values_per_pid)
    write_csv_file(name, header, big_order_indexes, array2d)
    if is_output_excel:
        write_excel_file(name, header, big_order_indexes, array2d)
    if is_view_graph:
        view_line_graph(id, name, header, big_order_indexes, array2d)


def add_time_and_value_array2d(time: str, pids: List[str], mem_dict: Dict, array2d: List[List[str]]):
    """

    :param time:
    :param pids:
    :param mem_dict:
    :param array2d:
    :return: void
    """
    if time == '' or len(mem_dict) == 0:
        return
    record: List[str] = [''] * (len(pids) + 1)
    record[0] = time
    for pid_key in mem_dict.keys():
        record[pids.index(pid_key) + 1] = mem_dict[pid_key]
    array2d.append(record)  # array2d record : time + pids


def add_time_and_mem_cpu_array2d(time: str, mem_pids: List[str], cpu_pids: List[str], mem_dict: Dict, cpu_dict: Dict,
                                 time_mem_array2d: List[List[str]], time_cpu_array2d: List[List[str]]):
    """

    :param time:
    :param mem_pids:
    :param cpu_pids:
    :param mem_dict:
    :param cpu_dict:
    :param time_mem_array2d:
    :param time_cpu_array2d:
    :return: void
    """
    add_time_and_value_array2d(time, mem_pids, mem_dict, time_mem_array2d)
    add_time_and_value_array2d(time, cpu_pids, cpu_dict, time_cpu_array2d)


def analyze_pid_value_line(line: str, pids: List[str], value_dict: Dict, value_index: int):
    """

    :param line:
    :param pids:
    :param value_dict:
    :param value_index:
    :return: void
    """
    line_columns: List[str] = line.split()
    pid: str = line_columns[PID_INDEX] + '(' + line_columns[COMMAND_INDEX] + ')'
    mem: str = line_columns[value_index]
    if float(mem) >= FILTER_VALUE:
        if pid not in pids:
            pids.append(pid)
        value_dict[pid] = mem


def analyze_top_log_lines(lines: List[str], mem_pids: List[str], cpu_pids: List[str], time_mem_array2d: List[List[str]],
                          time_cpu_array2d: List[List[str]]):
    """

    :param lines:
    :param mem_pids:
    :param cpu_pids:
    :param time_mem_array2d:
    :param time_cpu_array2d:
    :return: void
    """
    is_pid_value_block = False
    mem_dict: Dict = {}
    cpu_dict: Dict = {}
    time = ''
    for line in lines:
        line_columns = line.split();
        if line.startswith('top -'):
            time = line_columns[2]
            add_time_and_mem_cpu_array2d(time, mem_pids, cpu_pids, mem_dict, cpu_dict, time_mem_array2d,
                                         time_cpu_array2d)
            is_pid_value_block = False
            mem_dict = {}
            cpu_dict = {}
            continue
        if line.startswith('    PID'):
            is_pid_value_block = True
            continue
        if is_pid_value_block and len(line_columns) == PID_VALUE_COLUMN_COUNT:
            analyze_pid_value_line(line, mem_pids, mem_dict, MEM_INDEX)
            analyze_pid_value_line(line, cpu_pids, cpu_dict, CPU_INDEX)
    add_time_and_mem_cpu_array2d(time, mem_pids, cpu_pids, mem_dict, cpu_dict, time_mem_array2d, time_cpu_array2d)


def analyze_top_log(file_path: str, is_output_excel: bool, is_view_graph: bool):
    """

    :param file_path:
    :param is_output_excel:
    :return: void
    """
    mem_pids: List[str] = []
    cpu_pids: List[str] = []
    time_mem_array2d: List[List[str]] = []
    time_cpu_array2d: List[List[str]] = []
    with open(file_path, 'r') as f:
        lines = f.readlines()
        analyze_top_log_lines(lines, mem_pids, cpu_pids, time_mem_array2d, time_cpu_array2d)
    write_file_and_view_graph(0, OUTPUT_TOP_MEM_NAME, mem_pids, time_mem_array2d, is_output_excel, is_view_graph)
    write_file_and_view_graph(1, OUTPUT_TOP_CPU_NAME, cpu_pids, time_cpu_array2d, is_output_excel, is_view_graph)
    if is_view_graph:
        plt.show()


def main(args: List[str]):
    """

    :param args:
    :return: void
    """
    is_output_excel = False
    is_view_graph = False
    if EXCEL_OPTION in args:
        is_output_excel = True
    if VIEW_GRAPH_OPTION in args:
        is_view_graph = True
    file_paths = glob.glob("../input/top_*.log")
    for file_path in file_paths:
        analyze_top_log(file_path, is_output_excel, is_view_graph)


main(sys.argv)
